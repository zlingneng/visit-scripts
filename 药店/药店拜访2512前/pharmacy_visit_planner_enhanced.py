#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强版药店拜访规划系统
基于百度地图自动规划药店拜访路线，生成Excel拜访计划
功能：
1. 从指定药店开始搜索附近药店
2. 智能时间规划（考虑午休、工作时间限制）
3. 距离计算和路程时间估算
4. 生成详细的Excel拜访计划
5. 自动点击药店获取详情页URL
"""

import time
import random
import re
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class EnhancedPharmacyVisitPlanner:
    def __init__(self, start_pharmacy_name="一心堂大健康药店(花果园C区店)"):
        self.driver = None
        self.wait = None
        self.start_pharmacy_name = start_pharmacy_name
        self.visited_pharmacies = set()  # 已规划拜访的药店唯一标识集合（名称+地址）
        self.visit_plan = []  # 拜访计划列表
        self.current_time = None  # 当前规划时间
        self.max_pharmacies = random.randint(16, 19)  # 一天随机规划16-19家药店
        self.setup_driver()
        
    def setup_driver(self):
        """设置Chrome浏览器驱动"""
        chrome_options = Options()
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 10)
            print("Chrome浏览器启动成功")
        except Exception as e:
            print(f"浏览器启动失败: {e}")
            raise
    
    def init_visit_time(self):
        """初始化拜访开始时间（早上8:30-9:00随机）"""
        today = datetime.now().replace(hour=8, minute=30, second=0, microsecond=0)
        random_minutes = random.randint(0, 30)  # 8:30-9:00之间随机
        self.current_time = today + timedelta(minutes=random_minutes)
        print(f"拜访开始时间: {self.current_time.strftime('%H:%M')}")
    
    def calculate_travel_time(self, distance_str):
        """根据距离计算路程时间"""
        try:
            # 提取距离数字
            distance_match = re.search(r'(\d+(?:\.\d+)?)([km米])', distance_str)
            if not distance_match:
                return 5  # 默认5分钟
            
            distance_value = float(distance_match.group(1))
            unit = distance_match.group(2)
            
            # 转换为米
            if unit == 'km':
                distance_meters = distance_value * 1000
            elif unit == 'k':
                distance_meters = distance_value * 1000
            else:
                distance_meters = distance_value
            
            # 计算时间（分钟）
            if distance_meters <= 2000:  # 500米内步行
                travel_time = max(2, int(distance_meters / 80)) + random.randint(2, 5)  # 步行速度约80米/分钟，额外增加2-5分钟随机时间
            elif distance_meters <= 5000:  # 2公里内骑行或短途交通
                travel_time = max(5, int(distance_meters / 200)) + random.randint(5, 10)  # 约200米/分钟
            else:  # 长距离交通
                travel_time = max(10, int(distance_meters / 300)) + random.randint(10, 20)  # 约300米/分钟
            
            return travel_time
        except:
            return 5  # 默认5分钟
    
    def is_lunch_time(self, check_time):
        """检查是否是午休时间（12:00-13:30）"""
        lunch_start = check_time.replace(hour=12, minute=0)
        lunch_end = check_time.replace(hour=13, minute=random.randint(29, 46))
        return lunch_start <= check_time <= lunch_end
    
    def is_work_time_over(self, check_time):
        """检查是否超过工作时间（18:30）"""
        work_end = check_time.replace(hour=18, minute=30)
        return check_time > work_end
    
    def adjust_time_for_lunch(self, planned_time):
        """调整午休时间"""
        if self.is_lunch_time(planned_time):
            # 如果在午休时间，调整到13:30之后
            lunch_end = planned_time.replace(hour=13, minute=random.randint(29, 46))
            return lunch_end
        return planned_time
    
    def navigate_to_pharmacy_detail(self, pharmacy_name):
        """导航到药店详情页并获取URL"""
        try:
            # 打开百度地图
            self.driver.get('https://map.baidu.com')
            time.sleep(3)
            
            # 搜索药店
            search_box = self.wait.until(
                EC.presence_of_element_located((By.ID, "sole-input"))
            )
            search_box.clear()
            search_box.send_keys(pharmacy_name)
            
            # 点击搜索
            search_button = self.wait.until(
                EC.element_to_be_clickable((By.ID, "search-button"))
            )
            search_button.click()
            time.sleep(2)
            
            # 点击搜索结果
            result_link = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, f"//a[contains(text(), '{pharmacy_name}')]"))
            )
            result_link.click()
            time.sleep(2)
            
            # 获取当前URL
            current_url = self.driver.current_url
            print(f"获取到药店详情页URL: {current_url}")
            return current_url
            
        except Exception as e:
            print(f"导航到药店详情页失败: {e}")
            return ""
    
    def navigate_to_url(self, url):
        """直接导航到指定URL"""
        try:
            self.driver.get(url)
            time.sleep(3)
            print(f"已导航到: {url}")
            return True
        except Exception as e:
            print(f"导航到URL失败: {e}")
            return False
    
    def get_pharmacy_unique_key(self, name, address):
        """生成药店唯一标识（名称+地址）"""
        return f"{name.strip()}|{address.strip()}"
    
    def parse_distance_for_sorting(self, distance_str):
        """解析距离字符串用于排序，返回以米为单位的数值"""
        try:
            # 提取距离数字
            distance_match = re.search(r'(\d+(?:\.\d+)?)([km米])', distance_str)
            if not distance_match:
                return 999999  # 无法解析的距离放到最后
            
            distance_value = float(distance_match.group(1))
            unit = distance_match.group(2)
            
            # 转换为米
            if unit == 'km':
                return distance_value * 1000
            elif unit == 'k':
                return distance_value * 1000
            else:
                return distance_value
        except:
            return 999999
    
    def get_nearby_pharmacies(self, max_pages=3):
        """获取附近药店信息，支持翻页"""
        try:
            # 点击附近按钮
            nearby_button = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//span[@class='buttons-nearby-text' and text()='附近']"))
            )
            nearby_button.click()
            time.sleep(1)
            
            # 在附近搜索框输入"药店"
            nearby_input = self.wait.until(
                EC.presence_of_element_located((By.ID, "nearby-input"))
            )
            nearby_input.clear()
            nearby_input.send_keys("药店")
            time.sleep(1)
            
            # 点击附近搜索按钮
            search_button = self.wait.until(
                EC.element_to_be_clickable((By.ID, "search-button"))
            )
            search_button.click()
            time.sleep(3)
            
            all_pharmacies = []
            current_page = 1
            
            # 循环处理多页结果
            while current_page <= max_pages:
                print(f"\n正在处理第{current_page}页搜索结果...")
                page_pharmacies = []
                
                # 获取当前页面的药店列表
                for i in range(1, 21):  # 获取更多结果用于排序
                    try:
                        li_element = self.driver.find_element(By.CSS_SELECTOR, f'li[data-index="{i}"]')
                        
                        # 提取名称
                        try:
                            name_element = li_element.find_element(By.CSS_SELECTOR, 'a[data-stat-code="poisearch.all.title"]')
                            name = name_element.text.strip()
                        except:
                            name = "未找到名称"
                        
                        # 提取地址
                        try:
                            addr_element = li_element.find_element(By.CSS_SELECTOR, '.addr .n-grey')
                            address = addr_element.get_attribute('title') or addr_element.text.strip()
                        except:
                            address = "未找到地址"
                        
                        # 提取距离
                        try:
                            distance_element = li_element.find_element(By.CSS_SELECTOR, '.mt_5.h_20 .n-grey')
                            distance = distance_element.text.strip()
                        except:
                            distance = "未找到距离"
                        
                        # 生成唯一标识
                        unique_key = self.get_pharmacy_unique_key(name, address)
                        
                        # 检查是否已被规划拜访（使用唯一标识）
                        if unique_key not in self.visited_pharmacies and name != "未找到名称":
                            pharmacy_info = {
                                'index': i,
                                'name': name,
                                'address': address,
                                'distance': distance,
                                'distance_meters': self.parse_distance_for_sorting(distance),
                                'unique_key': unique_key,
                                'element': li_element,
                                'page': current_page
                            }
                            page_pharmacies.append(pharmacy_info)
                            print(f"发现附近药店: {name} - {address} ({distance})")
                        else:
                            print(f"跳过已规划药店: {name} - {address}")
                        
                    except Exception as e:
                        # 如果找不到更多元素，跳出循环
                        if "no such element" in str(e).lower():
                            break
                        print(f"提取第{i}个药店信息时出错: {e}")
                        continue
                
                # 将当前页面的药店添加到总列表
                all_pharmacies.extend(page_pharmacies)
                
                # 如果当前页面有未访问的药店，或者已经是最后一页，就不再翻页
                if page_pharmacies or current_page >= max_pages:
                    break
                
                # 尝试点击下一页
                try:
                    next_page_link = self.driver.find_element(By.CSS_SELECTOR, 'a[tid="toNextPage"]')
                    if next_page_link.is_enabled():
                        print(f"当前页面没有未访问药店，点击下一页...")
                        next_page_link.click()
                        time.sleep(3)
                        current_page += 1
                    else:
                        print("已到达最后一页")
                        break
                except Exception as e:
                    print(f"无法找到下一页按钮或已到达最后一页: {e}")
                    break
            
            # 按距离排序，选择最近的未访问药店
            if all_pharmacies:
                sorted_pharmacies = sorted(all_pharmacies, key=lambda x: x['distance_meters'])
                print(f"\n按距离排序后的药店列表（共{len(sorted_pharmacies)}家）:")
                for i, pharmacy in enumerate(sorted_pharmacies[:5]):  # 显示前5个
                    print(f"  {i+1}. {pharmacy['name']} - {pharmacy['distance']} (第{pharmacy['page']}页)")
                return sorted_pharmacies
            else:
                print("没有找到未访问的附近药店")
                return []
            
        except Exception as e:
            print(f"获取附近药店信息失败: {e}")
            return []
    
    def click_pharmacy_and_get_url(self, pharmacy_info):
        """点击药店并获取详情页URL"""
        try:
            # 点击药店名称链接
            name_link = pharmacy_info['element'].find_element(By.CSS_SELECTOR, 'a[data-stat-code="poisearch.all.title"]')
            name_link.click()
            time.sleep(2)
            
            # 获取详情页URL
            detail_url = self.driver.current_url
            print(f"获取到药店详情页URL: {detail_url}")
            return detail_url
            
        except Exception as e:
            print(f"点击药店获取URL失败: {e}")
            return ""
    
    def add_pharmacy_to_plan(self, pharmacy_name, address, detail_url, distance=""):
        """添加药店到拜访计划"""
        # 计算拜访时长（10-20分钟随机）
        visit_duration = random.randint(10, 20)
        
        # 计算路程时间
        travel_time = self.calculate_travel_time(distance) if distance else 0
        
        # 如果不是第一家药店，需要加上路程时间
        if len(self.visit_plan) > 0:
            self.current_time += timedelta(minutes=travel_time)
        
        # 调整午休时间
        self.current_time = self.adjust_time_for_lunch(self.current_time)
        
        # 检查是否超过工作时间
        if self.is_work_time_over(self.current_time):
            print(f"已超过工作时间18:30，停止规划")
            return False
        
        # 计算拜访结束时间
        visit_start = self.current_time
        visit_end = visit_start + timedelta(minutes=visit_duration)
        
        # 添加到计划
        plan_item = {
            '编号': len(self.visit_plan) + 1,
            '药店名称': pharmacy_name,
            '地址': address,
            '拜访开始时间': visit_start.strftime('%H:%M'),
            '拜访结束时间': visit_end.strftime('%H:%M'),
            '药店详情页URL': detail_url,
            '距离': distance,
            '路程时间(分钟)': travel_time if len(self.visit_plan) > 0 else 0
        }
        
        self.visit_plan.append(plan_item)
        # 使用唯一标识（名称+地址）记录已访问药店
        unique_key = self.get_pharmacy_unique_key(pharmacy_name, address)
        self.visited_pharmacies.add(unique_key)
        
        # 更新当前时间到拜访结束时间
        self.current_time = visit_end
        
        print(f"已添加到计划: {pharmacy_name} ({visit_start.strftime('%H:%M')}-{visit_end.strftime('%H:%M')})")
        return True
    
    def create_visit_plan(self):
        """创建完整的拜访计划"""
        print("开始创建药店拜访计划...")
        
        # 初始化时间
        self.init_visit_time()
        
        # 第一步：搜索起始药店并添加到计划
        print(f"\n=== 搜索起始药店: {self.start_pharmacy_name} ===")
        start_url = self.navigate_to_pharmacy_detail(self.start_pharmacy_name)
        if not start_url:
            print("无法找到起始药店，退出规划")
            return
        
        # 添加起始药店到计划（需要获取真实地址）
        try:
            # 尝试获取起始药店的真实地址
            addr_elements = self.driver.find_elements(By.CSS_SELECTOR, '.row.addr .n-grey')
            start_address = "贵阳市"  # 默认地址
            if addr_elements:
                # 优先使用title属性，如果没有则使用text内容
                start_address = addr_elements[0].get_attribute('title') or addr_elements[0].text.strip() or "贵阳市"
        except:
            start_address = "贵阳市"
        
        self.add_pharmacy_to_plan(self.start_pharmacy_name, start_address, start_url)
        
        # 循环搜索附近药店
        while len(self.visit_plan) < self.max_pharmacies:
            print(f"\n=== 第{len(self.visit_plan)}轮搜索附近药店 ===")
            
            # 使用最后一条数据的URL访问浏览器
            if len(self.visit_plan) > 1:
                last_url = self.visit_plan[-1]['药店详情页URL']
                print(f"使用最后一条数据的URL: {last_url}")
                if not self.navigate_to_url(last_url):
                    print("无法访问最后一条数据的URL，退出规划")
                    break
            
            # 获取附近药店
            nearby_pharmacies = self.get_nearby_pharmacies()
            if not nearby_pharmacies:
                print("没有找到更多附近药店，结束规划")
                break
            
            # 选择距离最近的未访问药店（已按距离排序）
            selected_pharmacy = nearby_pharmacies[0]  # 第一个就是距离最近的
            print(f"选择距离最近的药店: {selected_pharmacy['name']} ({selected_pharmacy['distance']})")
            
            # 点击药店获取详情页URL
            detail_url = self.click_pharmacy_and_get_url(selected_pharmacy)
            if not detail_url:
                print("无法获取药店详情页URL，跳过")
                continue
            
            # 添加到计划
            success = self.add_pharmacy_to_plan(
                selected_pharmacy['name'],
                selected_pharmacy['address'],
                detail_url,
                selected_pharmacy['distance']
            )
            
            if not success:
                break
            
            # 短暂等待
            time.sleep(1)
        
        print(f"\n拜访计划创建完成！共规划 {len(self.visit_plan)} 家药店")
    
    def save_to_excel(self, filename=None):
        """保存拜访计划到Excel文件"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"贵阳药店拜访计划_增强版_{timestamp}.xlsx"
        
        # 创建DataFrame
        df = pd.DataFrame(self.visit_plan)
        
        # 保存到Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='拜访计划', index=False)
            
            # 获取工作表进行格式化
            worksheet = writer.sheets['拜访计划']
            
            # 设置列宽
            column_widths = {
                'A': 8,   # 编号
                'B': 25,  # 药店名称
                'C': 30,  # 地址
                'D': 12,  # 拜访开始时间
                'E': 12,  # 拜访结束时间
                'F': 50,  # 药店详情页URL
                'G': 10,  # 距离
                'H': 12   # 路程时间
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置标题行格式
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置数据行格式
            for row in worksheet.iter_rows(min_row=2, max_row=len(self.visit_plan)+1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"拜访计划已保存到: {filename}")
        return filename
    
    def close(self):
        """关闭浏览器"""
        if self.driver:
            self.driver.quit()
            print("浏览器已关闭")
    
    def run(self, start_pharmacy=None):
        """运行完整的拜访规划流程"""
        if start_pharmacy:
            self.start_pharmacy_name = start_pharmacy
        
        try:
            # 创建拜访计划
            self.create_visit_plan()
            
            # 保存到Excel
            if self.visit_plan:
                filename = self.save_to_excel()
                
                # 显示计划摘要
                print("\n=== 拜访计划摘要 ===")
                for item in self.visit_plan:
                    print(f"{item['编号']}. {item['药店名称']} ({item['拜访开始时间']}-{item['拜访结束时间']}) - {item['距离']}")
                
                return filename
            else:
                print("没有生成任何拜访计划")
                return None
                
        except Exception as e:
            print(f"运行过程中出现错误: {e}")
            return None
        finally:
            # 保持浏览器打开以便观察
            input("\n按回车键关闭浏览器...")
            self.close()

def main():
    """主函数"""
    print("=== 增强版药店拜访规划系统 ===")
    print("功能特点：")
    print("1. 从指定药店开始搜索附近最近的药店")
    print("2. 智能时间规划（考虑午休12:00-13:30，晚上18:30结束）")
    print("3. 根据距离计算路程时间")
    print("4. 每次用最后一条数据的URL进行下一轮搜索")
    print("5. 按距离排序选择最近的未访问药店")
    print("6. 使用药店名称+地址确保唯一性，避免重复规划")
    print("7. 生成详细的Excel拜访计划")
    print("8. 一天最多规划20家药店")
    
    # 可以修改起始药店
    start_pharmacy = input("\n请输入起始药店名称（直接回车使用默认）: ").strip()
    if not start_pharmacy:
        start_pharmacy = "一心堂大健康药店(花果园C区店)"
    
    print(f"\n开始从 '{start_pharmacy}' 规划拜访路线...")
    
    planner = EnhancedPharmacyVisitPlanner(start_pharmacy)
    result_file = planner.run()
    
    if result_file:
        print(f"\n✅ 拜访计划已生成: {result_file}")
    else:
        print("\n❌ 拜访计划生成失败")

if __name__ == "__main__":
    main()