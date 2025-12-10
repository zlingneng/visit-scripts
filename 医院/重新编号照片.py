import os
import shutil
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

def renumber_photos_and_update_excel():
    """
    按照照片名称排序后重新编号，更新照片名称，并在Excel中添加新拜访编号列
    """
    # 设置文件路径
    excel_path = '/Users/a000/Documents/济生/医院拜访25/贵州医院拜访250115-5/贵州医院拜访250115-5_修改后.xlsx'
    image_source_dir = '/Users/a000/Pictures/2501-5/科室拜访'
    image_dest_dir = '/Users/a000/Documents/济生/医院拜访25/贵州医院拜访250115-5/科室'
    
    # 创建目标图片目录（如果不存在）
    Path(image_dest_dir).mkdir(parents=True, exist_ok=True)
    
    # 获取所有图片文件并按名称排序
    image_files = []
    if os.path.exists(image_source_dir):
        for filename in os.listdir(image_source_dir):
            if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                image_files.append(filename)
    
    # 按文件名转化为数值排序（文件名都是数字字符）
    def numeric_sort_key(filename):
        name_without_ext = os.path.splitext(filename)[0]
        try:
            return int(name_without_ext)  # 直接将文件名转换为数字
        except ValueError:
            return float('inf')  # 非数字文件名排在最后
    
    image_files.sort(key=numeric_sort_key)
    
    print(f"找到 {len(image_files)} 个图片文件")
    
    # 创建原照片名称到新编号的映射
    photo_mapping = {}
    for i, filename in enumerate(image_files, 1):
        photo_name_without_ext = os.path.splitext(filename)[0]
        photo_mapping[photo_name_without_ext] = i
        print(f"原照片名: {photo_name_without_ext} -> 新编号: {i}")
    
    # 读取Excel文件
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook['科室']
        
        # 添加新拜访编号列（如果不存在）
        if '新拜访编号' not in [cell.value for cell in sheet[1]]:
            # 找到最后一列
            max_col = sheet.max_column
            sheet.cell(row=1, column=max_col + 1, value='新拜访编号')
            new_visit_col = max_col + 1
        else:
            # 找到新拜访编号列的位置
            new_visit_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == '新拜访编号':
                    new_visit_col = col
                    break
        
        # 更新Excel中的新拜访编号
        for row in range(2, sheet.max_row + 1):
            # 假设第一列是拜访编号或照片名称
            original_name = str(sheet.cell(row=row, column=1).value) if sheet.cell(row=row, column=1).value else ""
            
            if original_name in photo_mapping:
                new_number = photo_mapping[original_name]
                sheet.cell(row=row, column=new_visit_col, value=new_number)
                print(f"Excel行 {row}: {original_name} -> 新编号 {new_number}")
            else:
                # 没有关联上的数据行，新拜访编号为空
                sheet.cell(row=row, column=new_visit_col, value=None)
                print(f"Excel行 {row}: {original_name} -> 无对应照片")
        
        # 保存Excel文件
        workbook.save(excel_path)
        print(f"Excel文件已更新: {excel_path}")
        
    except Exception as e:
        print(f"处理Excel文件时出错: {e}")
        return
    
    # 重命名并复制图片文件
    for i, filename in enumerate(image_files, 1):
        try:
            src_path = os.path.join(image_source_dir, filename)
            file_ext = os.path.splitext(filename)[1]
            new_filename = f"{i}{file_ext}"
            dest_path = os.path.join(image_dest_dir, new_filename)
            
            # 复制并重命名文件
            shutil.copy2(src_path, dest_path)
            print(f"已复制并重命名: {filename} -> {new_filename}")
            
        except Exception as e:
            print(f"处理图片 {filename} 时出错: {e}")
    
    print("\n重编号完成！")
    print(f"- 处理了 {len(image_files)} 个图片文件")
    print(f"- 图片已保存到: {image_dest_dir}")
    print(f"- Excel文件已更新: {excel_path}")

if __name__ == "__main__":
    renumber_photos_and_update_excel()